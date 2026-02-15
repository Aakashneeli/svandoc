import json
import unittest
from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook

from svandoc_backend.export_service import build_csv_export, build_json_export, build_xlsx_export


class ExportServiceTests(unittest.TestCase):
    def test_build_json_export_returns_schema_compatible_payload(self) -> None:
        payload = {
            "schema_version": "1.0.0",
            "document_type": "invoice",
            "metadata": {"document_id": "doc-1", "source_file_name": "invoice.pdf", "page_count": 1},
            "vendor": {"name": "ACME Inc", "tax_id": None, "address": None, "email": None},
            "customer": None,
            "invoice": {
                "invoice_number": "INV-1",
                "issue_date": "2026-02-15",
                "due_date": None,
                "purchase_order_number": None,
            },
            "amounts": {
                "currency": "USD",
                "subtotal": 100.0,
                "tax": 8.75,
                "shipping": None,
                "discount": None,
                "total": 108.75,
            },
            "line_items": [],
            "payment_terms": None,
            "confidence": {"overall": 0.95, "fields": {"amounts.total": 0.97}},
            "raw_text": "INVOICE OCR RAW",
            "review_required": False,
            "warnings": [],
        }

        exported = build_json_export(payload)
        exported_payload = json.loads(exported.decode("utf-8"))
        self.assertEqual(exported_payload["schema_version"], "1.0.0")
        self.assertEqual(exported_payload["document_type"], "invoice")
        self.assertEqual(exported_payload["amounts"]["total"], 108.75)

    def test_build_json_export_raises_for_non_canonical_payload(self) -> None:
        with self.assertRaises(ValueError):
            build_json_export(
                {
                    "schema_version": "1.0.0",
                    "document_type": "invoice",
                    "metadata": {"document_id": "doc-1"},
                }
            )

    def test_build_csv_export_produces_deterministic_headers_and_values(self) -> None:
        payload = {
            "schema_version": "1.0.0",
            "document_type": "invoice",
            "metadata": {"document_id": "doc-1", "source_file_name": "invoice.pdf", "page_count": 1},
            "vendor": {"name": "ACME Inc", "tax_id": None, "address": "NY", "email": None},
            "customer": {"name": "Buyer LLC", "address": None},
            "invoice": {
                "invoice_number": "INV-1",
                "issue_date": "2026-02-15",
                "due_date": None,
                "purchase_order_number": None,
            },
            "amounts": {
                "currency": "USD",
                "subtotal": 100.0,
                "tax": 8.75,
                "shipping": None,
                "discount": None,
                "total": 108.75,
            },
            "line_items": [{"description": "Service Fee", "quantity": 1, "unit_price": 108.75, "line_total": 108.75}],
            "payment_terms": "NET-15",
            "confidence": {"overall": 0.95, "fields": {"amounts.total": 0.97}},
            "raw_text": "INVOICE OCR RAW",
            "review_required": False,
            "warnings": [],
        }

        first = build_csv_export(payload).decode("utf-8")
        second = build_csv_export(payload).decode("utf-8")

        self.assertEqual(first, second)
        lines = first.strip().split("\n")
        self.assertEqual(len(lines), 2)
        self.assertEqual(
            lines[0],
            "document_id,source_file_name,page_count,schema_version,document_type,vendor_name,vendor_tax_id,vendor_address,vendor_email,customer_name,customer_address,invoice_number,issue_date,due_date,purchase_order_number,currency,subtotal,tax,shipping,discount,total,payment_terms,review_required,warning_count,line_item_count,line_items_json,confidence_overall",
        )
        self.assertIn("doc-1,invoice.pdf,1,1.0.0,invoice,ACME Inc,,NY,,Buyer LLC,,INV-1,2026-02-15,,,USD,100,8.75,,,108.75,NET-15,false,0,1,", lines[1])
        self.assertIn("\"[{\"\"description\"\":\"\"Service Fee\"\",\"\"line_total\"\":108.75,\"\"quantity\"\":1,\"\"unit_price\"\":108.75}]\",0.95", lines[1])

    def test_build_csv_export_raises_for_non_canonical_payload(self) -> None:
        with self.assertRaises(ValueError):
            build_csv_export(
                {
                    "schema_version": "1.0.0",
                    "document_type": "invoice",
                    "metadata": {"document_id": "doc-1"},
                }
            )

    def test_build_xlsx_export_preserves_numeric_and_date_types(self) -> None:
        payload = {
            "schema_version": "1.0.0",
            "document_type": "invoice",
            "metadata": {"document_id": "doc-1", "source_file_name": "invoice.pdf", "page_count": 1},
            "vendor": {"name": "ACME Inc", "tax_id": None, "address": "NY", "email": None},
            "customer": {"name": "Buyer LLC", "address": None},
            "invoice": {
                "invoice_number": "INV-1",
                "issue_date": "2026-02-15",
                "due_date": "2026-03-01",
                "purchase_order_number": None,
            },
            "amounts": {
                "currency": "USD",
                "subtotal": 100.0,
                "tax": 8.75,
                "shipping": None,
                "discount": None,
                "total": 108.75,
            },
            "line_items": [{"description": "Service Fee", "quantity": 1, "unit_price": 108.75, "line_total": 108.75}],
            "payment_terms": "NET-15",
            "confidence": {"overall": 0.95, "fields": {"amounts.total": 0.97}},
            "raw_text": "INVOICE OCR RAW",
            "review_required": False,
            "warnings": [],
        }

        content = build_xlsx_export(payload)
        workbook = load_workbook(filename=BytesIO(content))
        summary = workbook["summary"]
        headers = [cell.value for cell in summary[1]]
        values = [cell.value for cell in summary[2]]
        row = dict(zip(headers, values))

        self.assertEqual(row["document_id"], "doc-1")
        self.assertIsInstance(row["subtotal"], (int, float))
        self.assertEqual(row["subtotal"], 100.0)
        self.assertIsInstance(row["tax"], (int, float))
        self.assertEqual(row["tax"], 8.75)
        self.assertIsInstance(row["total"], (int, float))
        self.assertEqual(row["total"], 108.75)
        self.assertIsInstance(row["issue_date"], datetime)
        self.assertEqual(row["issue_date"].date().isoformat(), "2026-02-15")
        self.assertIsInstance(row["due_date"], datetime)
        self.assertEqual(row["due_date"].date().isoformat(), "2026-03-01")
        self.assertIn("line_items", workbook.sheetnames)

    def test_build_xlsx_export_raises_for_non_canonical_payload(self) -> None:
        with self.assertRaises(ValueError):
            build_xlsx_export(
                {
                    "schema_version": "1.0.0",
                    "document_type": "invoice",
                    "metadata": {"document_id": "doc-1"},
                }
            )


if __name__ == "__main__":
    unittest.main()
