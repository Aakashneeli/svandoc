import json
import os
import shutil
import unittest
import zipfile
from io import BytesIO
from pathlib import Path
from unittest.mock import patch
from uuid import uuid4

import sqlalchemy as sa
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.orm import sessionmaker

from svandoc_backend.db import Base, get_db_session
from svandoc_backend.main import app
from svandoc_backend.cloud_connectors import CloudConnectorError, CloudUploadResult
from svandoc_backend.models.document import Document
from svandoc_backend.models.export_artifact import ExportArtifact
from svandoc_backend.models.extraction_result import ExtractionResult
from svandoc_backend.models.xero_sync_log import XeroSyncLog
from svandoc_backend.google_sheets_export import GoogleSheetsExportResult
from svandoc_backend.quickbooks_connector import QuickBooksConnectorError, QuickBooksExportResult
from svandoc_backend.xero_connector import XeroConnectorError, XeroExportResult, XeroSyncAttempt


class ExportEndpointTests(unittest.TestCase):
    def setUp(self) -> None:
        workspace_tmp = Path("tests_tmp")
        workspace_tmp.mkdir(parents=True, exist_ok=True)
        self.test_dir = workspace_tmp / f"export-endpoint-{uuid4().hex}"
        self.test_dir.mkdir(parents=True, exist_ok=False)
        self.storage_dir = self.test_dir / "storage"
        self.db_path = self.test_dir / "export-endpoint-test.db"
        self.engine = sa.create_engine(f"sqlite:///{self.db_path.as_posix()}")
        self.SessionTesting = sessionmaker(bind=self.engine, autoflush=False, autocommit=False, expire_on_commit=False)
        Base.metadata.create_all(self.engine)

        self.previous_queue_backend = os.environ.get("QUEUE_BACKEND")
        self.previous_storage = os.environ.get("LOCAL_STORAGE_PATH")
        self.previous_storage_backend = os.environ.get("STORAGE_BACKEND")
        os.environ["QUEUE_BACKEND"] = "disabled"
        os.environ["LOCAL_STORAGE_PATH"] = str(self.storage_dir)
        os.environ["STORAGE_BACKEND"] = "local"

        def override_db_session():
            session = self.SessionTesting()
            try:
                yield session
            finally:
                session.close()

        app.dependency_overrides[get_db_session] = override_db_session
        self.client = TestClient(app)

    def tearDown(self) -> None:
        app.dependency_overrides.clear()
        self.engine.dispose()
        if self.previous_queue_backend is None:
            os.environ.pop("QUEUE_BACKEND", None)
        else:
            os.environ["QUEUE_BACKEND"] = self.previous_queue_backend
        if self.previous_storage is None:
            os.environ.pop("LOCAL_STORAGE_PATH", None)
        else:
            os.environ["LOCAL_STORAGE_PATH"] = self.previous_storage
        if self.previous_storage_backend is None:
            os.environ.pop("STORAGE_BACKEND", None)
        else:
            os.environ["STORAGE_BACKEND"] = self.previous_storage_backend
        shutil.rmtree(self.test_dir, ignore_errors=True)

    def _insert_document_with_extraction(self, document_id: str) -> None:
        session = self.SessionTesting()
        try:
            session.add(
                Document(
                    id=document_id,
                    team_id="team-a",
                    uploaded_by="user-a",
                    filename="invoice.pdf",
                    mime_type="application/pdf",
                    checksum=f"checksum-{document_id}",
                    storage_uri=str(self.test_dir / "invoice.pdf"),
                    page_count=1,
                )
            )
            session.add(
                ExtractionResult(
                    id=f"ext-{document_id}",
                    document_id=document_id,
                    schema_version="1.0.0",
                    doc_type="invoice",
                    raw_ocr_text="INVOICE OCR RAW",
                    structured_payload={
                        "schema_version": "1.0.0",
                        "document_type": "invoice",
                        "metadata": {"document_id": document_id, "source_file_name": "invoice.pdf", "page_count": 1},
                        "vendor": {"name": "ACME Inc", "tax_id": None, "address": None, "email": None},
                        "customer": None,
                        "invoice": {
                            "invoice_number": "INV-1",
                            "issue_date": "2026-02-15",
                            "due_date": None,
                            "purchase_order_number": None,
                        },
                        "amounts": {
                            "currency": "USD",
                            "subtotal": 100.0,
                            "tax": 8.75,
                            "shipping": None,
                            "discount": None,
                            "total": 108.75,
                        },
                        "line_items": [
                            {"description": "Service Fee", "quantity": 1, "unit_price": 108.75, "line_total": 108.75}
                        ],
                        "payment_terms": None,
                        "confidence": {"overall": 0.95, "fields": {"amounts.total": 0.97}},
                        "raw_text": "INVOICE OCR RAW",
                        "review_required": False,
                        "warnings": [],
                    },
                    confidence_map={"overall": 0.95, "fields": {"amounts.total": 0.97}},
                    is_review_required=False,
                )
            )
            session.commit()
        finally:
            session.close()

    def test_export_json_persists_artifact_and_file(self) -> None:
        document_id = "doc-export-json"
        self._insert_document_with_extraction(document_id)

        with patch("svandoc_backend.main.deliver_webhook_event") as webhook_mock:
            response = self.client.post(
                f"/api/documents/{document_id}/export",
                headers={"x-user-id": "editor-a"},
                json={"format": "json"},
            )
        self.assertEqual(response.status_code, 200)
        data = response.json()["data"]
        self.assertEqual(data["format"], "json")
        self.assertEqual(data["document_id"], document_id)
        self.assertEqual(data["created_by"], "editor-a")
        webhook_mock.assert_called_once()
        _, kwargs = webhook_mock.call_args
        self.assertEqual(kwargs["event_type"], "export.created")
        self.assertEqual(kwargs["data"]["document_id"], document_id)

        session = self.SessionTesting()
        try:
            artifact = session.get(ExportArtifact, data["artifact_id"])
        finally:
            session.close()
        self.assertIsNotNone(artifact)
        assert artifact is not None
        self.assertEqual(artifact.format, "json")
        stored_content = Path(artifact.storage_uri).read_text(encoding="utf-8")
        rendered_payload = json.loads(stored_content)
        self.assertEqual(rendered_payload["schema_version"], "1.0.0")
        self.assertEqual(rendered_payload["document_type"], "invoice")

    def test_export_xlsx_persists_excel_file(self) -> None:
        document_id = "doc-export-xlsx"
        self._insert_document_with_extraction(document_id)

        response = self.client.post(
            f"/api/documents/{document_id}/export",
            json={"format": "xlsx"},
        )
        self.assertEqual(response.status_code, 200)
        artifact_id = response.json()["data"]["artifact_id"]

        session = self.SessionTesting()
        try:
            artifact = session.get(ExportArtifact, artifact_id)
        finally:
            session.close()
        self.assertIsNotNone(artifact)
        assert artifact is not None
        self.assertEqual(artifact.format, "xlsx")
        workbook = load_workbook(filename=BytesIO(Path(artifact.storage_uri).read_bytes()))
        self.assertIn("summary", workbook.sheetnames)
        self.assertIn("line_items", workbook.sheetnames)

    def test_export_rejects_invalid_format(self) -> None:
        document_id = "doc-export-invalid"
        self._insert_document_with_extraction(document_id)

        response = self.client.post(
            f"/api/documents/{document_id}/export",
            json={"format": "pdf"},
        )
        self.assertEqual(response.status_code, 400)
        payload = response.json()
        self.assertEqual(payload["status"], "error")
        self.assertEqual(payload["error"]["code"], "VALIDATION_ERROR")

    def test_export_gdrive_persists_cloud_artifact(self) -> None:
        document_id = "doc-export-gdrive"
        self._insert_document_with_extraction(document_id)

        with patch("svandoc_backend.main.upload_to_google_drive") as mocked_upload:
            mocked_upload.return_value = CloudUploadResult(
                provider="gdrive",
                remote_id="drive-file-1",
                storage_uri="gdrive://drive-file-1",
            )
            response = self.client.post(
                f"/api/documents/{document_id}/export",
                json={
                    "format": "gdrive",
                    "cloud_access_token": "token-value",
                    "cloud_folder": "folder-1",
                },
            )

        self.assertEqual(response.status_code, 200)
        data = response.json()["data"]
        self.assertEqual(data["format"], "gdrive")
        self.assertEqual(data["storage_uri"], "gdrive://drive-file-1")
        self.assertEqual(data["delivery_status"], "completed")

    def test_export_onedrive_persists_cloud_artifact(self) -> None:
        document_id = "doc-export-onedrive"
        self._insert_document_with_extraction(document_id)

        with patch("svandoc_backend.main.upload_to_onedrive") as mocked_upload:
            mocked_upload.return_value = CloudUploadResult(
                provider="onedrive",
                remote_id="onedrive-file-1",
                storage_uri="onedrive://onedrive-file-1",
            )
            response = self.client.post(
                f"/api/documents/{document_id}/export",
                json={
                    "format": "onedrive",
                    "cloud_access_token": "token-value",
                    "cloud_folder": "svandoc",
                },
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["data"]["format"], "onedrive")
        self.assertEqual(response.json()["data"]["delivery_status"], "completed")

    def test_export_dropbox_persists_cloud_artifact(self) -> None:
        document_id = "doc-export-dropbox"
        self._insert_document_with_extraction(document_id)

        with patch("svandoc_backend.main.upload_to_dropbox") as mocked_upload:
            mocked_upload.return_value = CloudUploadResult(
                provider="dropbox",
                remote_id="dropbox-file-1",
                storage_uri="dropbox://dropbox-file-1",
            )
            response = self.client.post(
                f"/api/documents/{document_id}/export",
                json={
                    "format": "dropbox",
                    "cloud_access_token": "token-value",
                },
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["data"]["format"], "dropbox")
        self.assertEqual(response.json()["data"]["delivery_status"], "completed")

    def test_export_cloud_connector_failure_persists_failed_status(self) -> None:
        document_id = "doc-export-cloud-failed"
        self._insert_document_with_extraction(document_id)

        with patch("svandoc_backend.main.upload_to_google_drive") as mocked_upload:
            with patch("svandoc_backend.main.deliver_webhook_event") as webhook_mock:
                mocked_upload.side_effect = CloudConnectorError("google_drive_upload_failed:403")
                response = self.client.post(
                    f"/api/documents/{document_id}/export",
                    json={
                        "format": "gdrive",
                        "cloud_access_token": "token-value",
                    },
                )

        self.assertEqual(response.status_code, 502)
        payload = response.json()
        self.assertEqual(payload["error"]["code"], "EXPORT_DELIVERY_FAILED")
        artifact_id = payload["error"]["details"]["artifact_id"]
        webhook_mock.assert_called_once()
        _, kwargs = webhook_mock.call_args
        self.assertEqual(kwargs["event_type"], "export.created")
        self.assertEqual(kwargs["data"]["delivery_status"], "failed")

        session = self.SessionTesting()
        try:
            artifact = session.get(ExportArtifact, artifact_id)
        finally:
            session.close()
        self.assertIsNotNone(artifact)
        assert artifact is not None
        self.assertEqual(artifact.delivery_status, "failed")
        self.assertEqual(artifact.storage_uri, "failed://gdrive")

    def test_export_cloud_connector_requires_access_token(self) -> None:
        document_id = "doc-export-cloud-validation"
        self._insert_document_with_extraction(document_id)

        response = self.client.post(
            f"/api/documents/{document_id}/export",
            json={"format": "onedrive"},
        )
        self.assertEqual(response.status_code, 400)
        payload = response.json()
        self.assertEqual(payload["error"]["code"], "VALIDATION_ERROR")

    def test_export_gsheets_persists_artifact_with_connector_uri(self) -> None:
        document_id = "doc-export-gsheets"
        self._insert_document_with_extraction(document_id)

        with patch("svandoc_backend.main.append_to_google_sheet") as mocked_append:
            mocked_append.return_value = GoogleSheetsExportResult(
                spreadsheet_id="spreadsheet-123",
                sheet_name="InvoiceExports",
                updated_range="InvoiceExports!A1:AA2",
                updated_rows=2,
            )
            response = self.client.post(
                f"/api/documents/{document_id}/export",
                json={
                    "format": "gsheets",
                    "google_spreadsheet_id": "spreadsheet-123",
                    "google_sheet_name": "InvoiceExports",
                    "google_access_token": "token-value",
                },
            )

        self.assertEqual(response.status_code, 200)
        data = response.json()["data"]
        self.assertEqual(data["format"], "gsheets")
        self.assertEqual(data["storage_uri"], "gsheets://spreadsheet-123/InvoiceExports")

        session = self.SessionTesting()
        try:
            artifact = session.get(ExportArtifact, data["artifact_id"])
        finally:
            session.close()
        self.assertIsNotNone(artifact)
        assert artifact is not None
        self.assertEqual(artifact.format, "gsheets")
        self.assertEqual(artifact.storage_uri, "gsheets://spreadsheet-123/InvoiceExports")

    def test_export_gsheets_requires_oauth_fields(self) -> None:
        document_id = "doc-export-gsheets-validation"
        self._insert_document_with_extraction(document_id)

        response = self.client.post(
            f"/api/documents/{document_id}/export",
            json={"format": "gsheets"},
        )
        self.assertEqual(response.status_code, 400)
        payload = response.json()
        self.assertEqual(payload["error"]["code"], "VALIDATION_ERROR")

    def test_export_quickbooks_persists_artifact(self) -> None:
        document_id = "doc-export-quickbooks"
        self._insert_document_with_extraction(document_id)

        with patch("svandoc_backend.main.export_to_quickbooks") as mocked_export:
            mocked_export.return_value = QuickBooksExportResult(
                realm_id="realm-1",
                resource_id="purchase-100",
                storage_uri="quickbooks://realm-1/purchase-100",
            )
            response = self.client.post(
                f"/api/documents/{document_id}/export",
                json={
                    "format": "quickbooks",
                    "quickbooks_access_token": "token-value",
                    "quickbooks_realm_id": "realm-1",
                },
            )

        self.assertEqual(response.status_code, 200)
        data = response.json()["data"]
        self.assertEqual(data["format"], "quickbooks")
        self.assertEqual(data["storage_uri"], "quickbooks://realm-1/purchase-100")
        self.assertEqual(data["delivery_status"], "completed")

    def test_export_quickbooks_requires_required_fields(self) -> None:
        document_id = "doc-export-quickbooks-validation"
        self._insert_document_with_extraction(document_id)

        response = self.client.post(
            f"/api/documents/{document_id}/export",
            json={"format": "quickbooks"},
        )
        self.assertEqual(response.status_code, 400)
        payload = response.json()
        self.assertEqual(payload["error"]["code"], "VALIDATION_ERROR")

    def test_export_quickbooks_failure_persists_failed_status(self) -> None:
        document_id = "doc-export-quickbooks-failed"
        self._insert_document_with_extraction(document_id)

        with patch("svandoc_backend.main.export_to_quickbooks") as mocked_export:
            mocked_export.side_effect = QuickBooksConnectorError("quickbooks_export_failed:401")
            response = self.client.post(
                f"/api/documents/{document_id}/export",
                json={
                    "format": "quickbooks",
                    "quickbooks_access_token": "token-value",
                    "quickbooks_realm_id": "realm-1",
                },
            )

        self.assertEqual(response.status_code, 502)
        payload = response.json()
        self.assertEqual(payload["error"]["code"], "EXPORT_DELIVERY_FAILED")
        artifact_id = payload["error"]["details"]["artifact_id"]

        session = self.SessionTesting()
        try:
            artifact = session.get(ExportArtifact, artifact_id)
        finally:
            session.close()
        self.assertIsNotNone(artifact)
        assert artifact is not None
        self.assertEqual(artifact.delivery_status, "failed")
        self.assertEqual(artifact.storage_uri, "failed://quickbooks")

    def test_export_xero_persists_artifact_and_sync_logs(self) -> None:
        document_id = "doc-export-xero"
        self._insert_document_with_extraction(document_id)

        with patch("svandoc_backend.main.export_to_xero") as mocked_export:
            mocked_export.return_value = XeroExportResult(
                tenant_id="tenant-1",
                invoice_id="invoice-500",
                storage_uri="xero://tenant-1/invoice-500",
                attempts=[
                    XeroSyncAttempt(
                        attempt_number=1,
                        sync_status="synced",
                        status_code=200,
                        error_message=None,
                        external_reference="invoice-500",
                    )
                ],
            )
            response = self.client.post(
                f"/api/documents/{document_id}/export",
                json={
                    "format": "xero",
                    "xero_access_token": "token-value",
                    "xero_tenant_id": "tenant-1",
                },
            )

        self.assertEqual(response.status_code, 200)
        data = response.json()["data"]
        self.assertEqual(data["format"], "xero")
        self.assertEqual(data["storage_uri"], "xero://tenant-1/invoice-500")

        session = self.SessionTesting()
        try:
            logs = (
                session.query(XeroSyncLog)
                .filter(XeroSyncLog.document_id == document_id)
                .order_by(XeroSyncLog.attempt_number.asc())
                .all()
            )
        finally:
            session.close()
        self.assertEqual(len(logs), 1)
        self.assertEqual(logs[0].sync_status, "synced")
        self.assertEqual(logs[0].external_reference, "invoice-500")

    def test_export_xero_failure_persists_failed_artifact_and_retry_logs(self) -> None:
        document_id = "doc-export-xero-failed"
        self._insert_document_with_extraction(document_id)

        with patch("svandoc_backend.main.export_to_xero") as mocked_export:
            mocked_export.side_effect = XeroConnectorError(
                "xero_export_failed:http_status_429",
                attempts=[
                    XeroSyncAttempt(
                        attempt_number=1,
                        sync_status="retrying",
                        status_code=429,
                        error_message="http_status_429",
                    ),
                    XeroSyncAttempt(
                        attempt_number=2,
                        sync_status="failed",
                        status_code=429,
                        error_message="http_status_429",
                    ),
                ],
            )
            response = self.client.post(
                f"/api/documents/{document_id}/export",
                json={
                    "format": "xero",
                    "xero_access_token": "token-value",
                    "xero_tenant_id": "tenant-1",
                },
            )

        self.assertEqual(response.status_code, 502)
        payload = response.json()
        self.assertEqual(payload["error"]["code"], "EXPORT_DELIVERY_FAILED")
        artifact_id = payload["error"]["details"]["artifact_id"]

        session = self.SessionTesting()
        try:
            artifact = session.get(ExportArtifact, artifact_id)
            logs = (
                session.query(XeroSyncLog)
                .filter(XeroSyncLog.document_id == document_id)
                .order_by(XeroSyncLog.attempt_number.asc())
                .all()
            )
        finally:
            session.close()
        self.assertIsNotNone(artifact)
        assert artifact is not None
        self.assertEqual(artifact.delivery_status, "failed")
        self.assertEqual(artifact.storage_uri, "failed://xero")
        self.assertEqual(len(logs), 2)
        self.assertEqual(logs[0].sync_status, "retrying")
        self.assertEqual(logs[1].sync_status, "failed")

    def test_export_xero_requires_required_fields(self) -> None:
        document_id = "doc-export-xero-validation"
        self._insert_document_with_extraction(document_id)

        response = self.client.post(
            f"/api/documents/{document_id}/export",
            json={"format": "xero"},
        )
        self.assertEqual(response.status_code, 400)
        payload = response.json()
        self.assertEqual(payload["error"]["code"], "VALIDATION_ERROR")

    def test_export_sage_persists_strategy_artifact(self) -> None:
        document_id = "doc-export-sage"
        self._insert_document_with_extraction(document_id)

        response = self.client.post(
            f"/api/documents/{document_id}/export",
            json={"format": "sage"},
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()["data"]
        self.assertEqual(data["format"], "sage")
        session = self.SessionTesting()
        try:
            artifact = session.get(ExportArtifact, data["artifact_id"])
        finally:
            session.close()
        self.assertIsNotNone(artifact)
        assert artifact is not None
        plan = json.loads(Path(artifact.storage_uri).read_text(encoding="utf-8"))
        self.assertEqual(plan["provider"], "sage")
        self.assertEqual(plan["phases"][0]["phase"], "phase_1_file_exchange")

    def test_export_tally_persists_import_package(self) -> None:
        document_id = "doc-export-tally"
        self._insert_document_with_extraction(document_id)

        response = self.client.post(
            f"/api/documents/{document_id}/export",
            json={"format": "tally"},
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()["data"]
        self.assertEqual(data["format"], "tally")
        session = self.SessionTesting()
        try:
            artifact = session.get(ExportArtifact, data["artifact_id"])
        finally:
            session.close()
        self.assertIsNotNone(artifact)
        assert artifact is not None
        archive_bytes = Path(artifact.storage_uri).read_bytes()
        with zipfile.ZipFile(BytesIO(archive_bytes), mode="r") as archive:
            names = set(archive.namelist())
            self.assertIn("manifest.json", names)
            self.assertIn("voucher.xml", names)
            self.assertIn("summary.csv", names)

    def test_export_returns_404_when_document_missing(self) -> None:
        response = self.client.post(
            "/api/documents/missing-document/export",
            json={"format": "json"},
        )
        self.assertEqual(response.status_code, 404)
        payload = response.json()
        self.assertEqual(payload["error"]["code"], "DOCUMENT_NOT_FOUND")

    def test_export_returns_404_when_extraction_missing(self) -> None:
        document_id = "doc-export-no-extraction"
        session = self.SessionTesting()
        try:
            session.add(
                Document(
                    id=document_id,
                    team_id="team-a",
                    uploaded_by="user-a",
                    filename="invoice.pdf",
                    mime_type="application/pdf",
                    checksum=f"checksum-{document_id}",
                    storage_uri=str(self.test_dir / "invoice.pdf"),
                    page_count=1,
                )
            )
            session.commit()
        finally:
            session.close()

        response = self.client.post(
            f"/api/documents/{document_id}/export",
            json={"format": "json"},
        )
        self.assertEqual(response.status_code, 404)
        payload = response.json()
        self.assertEqual(payload["error"]["code"], "EXTRACTION_NOT_FOUND")


if __name__ == "__main__":
    unittest.main()
