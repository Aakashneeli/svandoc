import json
import os
import shutil
import unittest
from io import BytesIO
from pathlib import Path
from uuid import uuid4

import sqlalchemy as sa
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.orm import sessionmaker

from svandoc_backend.db import Base, get_db_session
from svandoc_backend.main import app
from svandoc_backend.models.document import Document
from svandoc_backend.models.export_artifact import ExportArtifact
from svandoc_backend.models.extraction_result import ExtractionResult
from svandoc_backend.models.job import Job


class EndToEndSmokeTests(unittest.TestCase):
    def setUp(self) -> None:
        workspace_tmp = Path("tests_tmp")
        workspace_tmp.mkdir(parents=True, exist_ok=True)
        self.test_dir = workspace_tmp / f"e2e-smoke-{uuid4().hex}"
        self.test_dir.mkdir(parents=True, exist_ok=False)
        self.storage_dir = self.test_dir / "storage"
        self.db_path = self.test_dir / "e2e-smoke.db"
        self.engine = sa.create_engine(f"sqlite:///{self.db_path.as_posix()}")
        self.SessionTesting = sessionmaker(bind=self.engine, autoflush=False, autocommit=False, expire_on_commit=False)
        Base.metadata.create_all(self.engine)

        self.previous_queue_backend = os.environ.get("QUEUE_BACKEND")
        self.previous_storage = os.environ.get("LOCAL_STORAGE_PATH")
        self.previous_storage_backend = os.environ.get("STORAGE_BACKEND")
        os.environ["QUEUE_BACKEND"] = "disabled"
        os.environ["LOCAL_STORAGE_PATH"] = str(self.storage_dir)
        os.environ["STORAGE_BACKEND"] = "local"

        def override_db_session():
            session = self.SessionTesting()
            try:
                yield session
            finally:
                session.close()

        app.dependency_overrides[get_db_session] = override_db_session
        self.client = TestClient(app)

    def tearDown(self) -> None:
        app.dependency_overrides.clear()
        self.engine.dispose()
        if self.previous_queue_backend is None:
            os.environ.pop("QUEUE_BACKEND", None)
        else:
            os.environ["QUEUE_BACKEND"] = self.previous_queue_backend
        if self.previous_storage is None:
            os.environ.pop("LOCAL_STORAGE_PATH", None)
        else:
            os.environ["LOCAL_STORAGE_PATH"] = self.previous_storage
        if self.previous_storage_backend is None:
            os.environ.pop("STORAGE_BACKEND", None)
        else:
            os.environ["STORAGE_BACKEND"] = self.previous_storage_backend
        shutil.rmtree(self.test_dir, ignore_errors=True)

    def _seed_extraction(self, document_id: str) -> None:
        session = self.SessionTesting()
        try:
            extraction = ExtractionResult(
                id=f"ext-{document_id}",
                document_id=document_id,
                schema_version="1.0.0",
                doc_type="invoice",
                raw_ocr_text="INVOICE OCR RAW",
                structured_payload={
                    "schema_version": "1.0.0",
                    "document_type": "invoice",
                    "metadata": {"document_id": document_id, "source_file_name": "invoice.pdf", "page_count": 1},
                    "vendor": {"name": "ACME Inc", "tax_id": None, "address": None, "email": None},
                    "customer": None,
                    "invoice": {
                        "invoice_number": "INV-1",
                        "issue_date": "2026-02-15",
                        "due_date": None,
                        "purchase_order_number": None,
                    },
                    "amounts": {
                        "currency": "USD",
                        "subtotal": 100.0,
                        "tax": 8.75,
                        "shipping": None,
                        "discount": None,
                        "total": 108.75,
                    },
                    "line_items": [
                        {"description": "Service Fee", "quantity": 1, "unit_price": 108.75, "line_total": 108.75}
                    ],
                    "payment_terms": None,
                    "confidence": {"overall": 0.95, "fields": {"amounts.total": 0.97, "vendor.name": 0.93}},
                    "raw_text": "INVOICE OCR RAW",
                    "review_required": False,
                    "warnings": [],
                },
                confidence_map={"overall": 0.95, "fields": {"amounts.total": 0.97, "vendor.name": 0.93}},
                is_review_required=False,
            )
            session.add(extraction)

            job = session.query(Job).filter(Job.document_id == document_id).one_or_none()
            if job is not None:
                job.status = "completed"

            session.commit()
        finally:
            session.close()

    def test_upload_review_export_smoke_flow(self) -> None:
        upload_response = self.client.post(
            "/api/documents/upload",
            headers={"x-team-id": "team-a", "x-user-id": "reviewer-a"},
            files=[("files", ("invoice.pdf", b"%PDF-1.7 smoke", "application/pdf"))],
        )
        self.assertEqual(upload_response.status_code, 200)
        upload_data = upload_response.json()["data"]
        document_id = upload_data["document_ids"][0]
        job_id = upload_data["job_ids"][0]

        session = self.SessionTesting()
        try:
            document = session.get(Document, document_id)
        finally:
            session.close()
        self.assertIsNotNone(document)
        assert document is not None
        self.assertTrue(Path(document.storage_uri).exists())

        job_response = self.client.get(f"/api/jobs/{job_id}")
        self.assertEqual(job_response.status_code, 200)
        self.assertEqual(job_response.json()["data"]["status"], "queued")

        self._seed_extraction(document_id)

        extraction_response = self.client.get(f"/api/documents/{document_id}/extraction")
        self.assertEqual(extraction_response.status_code, 200)
        self.assertEqual(extraction_response.json()["data"]["doc_type"], "invoice")
        self.assertEqual(extraction_response.json()["data"]["structured_payload"]["vendor"]["name"], "ACME Inc")

        patch_response = self.client.patch(
            f"/api/documents/{document_id}/extraction",
            headers={"x-user-id": "editor-a"},
            json={
                "corrections": [
                    {"field_path": "vendor.name", "new_value": "ACME Corporation"},
                    {"field_path": "amounts.total", "new_value": 110.25},
                ]
            },
        )
        self.assertEqual(patch_response.status_code, 200)
        patch_data = patch_response.json()["data"]
        self.assertEqual(patch_data["correction_count"], 2)
        self.assertEqual(patch_data["structured_payload"]["vendor"]["name"], "ACME Corporation")
        self.assertEqual(patch_data["structured_payload"]["amounts"]["total"], 110.25)

        exported_artifact_ids: list[str] = []
        for export_format in ("json", "csv", "xlsx"):
            export_response = self.client.post(
                f"/api/documents/{document_id}/export",
                headers={"x-user-id": "editor-a"},
                json={"format": export_format},
            )
            self.assertEqual(export_response.status_code, 200)
            export_data = export_response.json()["data"]
            self.assertEqual(export_data["format"], export_format)
            exported_artifact_ids.append(export_data["artifact_id"])

        session = self.SessionTesting()
        try:
            artifacts = [session.get(ExportArtifact, artifact_id) for artifact_id in exported_artifact_ids]
        finally:
            session.close()

        for artifact in artifacts:
            self.assertIsNotNone(artifact)
            assert artifact is not None
            self.assertTrue(Path(artifact.storage_uri).exists())

        json_artifact = next(artifact for artifact in artifacts if artifact is not None and artifact.format == "json")
        assert json_artifact is not None
        json_payload = json.loads(Path(json_artifact.storage_uri).read_text(encoding="utf-8"))
        self.assertEqual(json_payload["vendor"]["name"], "ACME Corporation")
        self.assertEqual(json_payload["amounts"]["total"], 110.25)

        csv_artifact = next(artifact for artifact in artifacts if artifact is not None and artifact.format == "csv")
        assert csv_artifact is not None
        csv_text = Path(csv_artifact.storage_uri).read_text(encoding="utf-8")
        self.assertIn("ACME Corporation", csv_text)
        self.assertIn("110.25", csv_text)

        xlsx_artifact = next(artifact for artifact in artifacts if artifact is not None and artifact.format == "xlsx")
        assert xlsx_artifact is not None
        workbook = load_workbook(filename=BytesIO(Path(xlsx_artifact.storage_uri).read_bytes()))
        self.assertIn("summary", workbook.sheetnames)
        self.assertIn("line_items", workbook.sheetnames)


if __name__ == "__main__":
    unittest.main()
